"""
Audit Logs router — admin-only.

GET  /audit-logs/        — paginated list with filters
GET  /audit-logs/export  — download as .xlsx (admin only)

Staff receive HTTP 403 on every endpoint here.
"""
from __future__ import annotations

import io
from datetime import datetime, date
from typing import List, Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.core.auth import require_admin
from app.database.dependencies import get_db
from app.models.audit_log import AuditLog
from app.models.user import User

router = APIRouter(prefix="/audit-logs", tags=["Audit Logs"])


# ── Response schema ────────────────────────────────────────────

class AuditLogResponse(BaseModel):
    id: int
    username: Optional[str]
    action: str
    module: str
    description: str
    created_at: datetime

    model_config = {"from_attributes": True}


class AuditLogPage(BaseModel):
    total: int
    page: int
    page_size: int
    items: List[AuditLogResponse]


# ── List (paginated + filtered) ────────────────────────────────

@router.get("/", response_model=AuditLogPage)
def list_audit_logs(
    # Filters
    username: Optional[str]   = Query(default=None),
    action: Optional[str]     = Query(default=None),
    module: Optional[str]     = Query(default=None),
    search: Optional[str]     = Query(default=None, description="Search in description"),
    date_from: Optional[date] = Query(default=None),
    date_to: Optional[date]   = Query(default=None),
    # Pagination
    page: int      = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=200),
    # Auth — ADMIN ONLY
    _admin: User   = Depends(require_admin),
    db: Session    = Depends(get_db),
):
    q = db.query(AuditLog)

    if username:
        q = q.filter(AuditLog.username.ilike(f"%{username}%"))
    if action:
        q = q.filter(AuditLog.action == action)
    if module:
        q = q.filter(AuditLog.module == module)
    if search:
        q = q.filter(AuditLog.description.ilike(f"%{search}%"))
    if date_from:
        q = q.filter(AuditLog.created_at >= datetime(date_from.year, date_from.month, date_from.day))
    if date_to:
        from datetime import timedelta
        q = q.filter(
            AuditLog.created_at < datetime(date_to.year, date_to.month, date_to.day) + timedelta(days=1)
        )

    total = q.count()
    items = (
        q.order_by(AuditLog.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    return AuditLogPage(
        total=total,
        page=page,
        page_size=page_size,
        items=items,
    )


# ── Export as Excel ────────────────────────────────────────────

@router.get("/export")
def export_audit_logs(
    username: Optional[str]   = Query(default=None),
    action: Optional[str]     = Query(default=None),
    module: Optional[str]     = Query(default=None),
    search: Optional[str]     = Query(default=None),
    date_from: Optional[date] = Query(default=None),
    date_to: Optional[date]   = Query(default=None),
    _admin: User               = Depends(require_admin),
    db: Session                = Depends(get_db),
):
    """Download audit logs as .xlsx — admin only."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail="openpyxl is not installed on the server")

    q = db.query(AuditLog)

    if username:
        q = q.filter(AuditLog.username.ilike(f"%{username}%"))
    if action:
        q = q.filter(AuditLog.action == action)
    if module:
        q = q.filter(AuditLog.module == module)
    if search:
        q = q.filter(AuditLog.description.ilike(f"%{search}%"))
    if date_from:
        q = q.filter(AuditLog.created_at >= datetime(date_from.year, date_from.month, date_from.day))
    if date_to:
        from datetime import timedelta
        q = q.filter(
            AuditLog.created_at < datetime(date_to.year, date_to.month, date_to.day) + timedelta(days=1)
        )

    logs = q.order_by(AuditLog.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Logs"

    headers = ["ID", "Date/Time", "Username", "Action", "Module", "Description"]
    header_fill = PatternFill(start_color="1E1B2E", end_color="1E1B2E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill  = header_fill
        cell.font  = header_font
        cell.alignment = Alignment(horizontal="center")

    col_widths = [8, 22, 20, 20, 18, 80]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    for row_idx, log in enumerate(logs, start=2):
        ws.cell(row=row_idx, column=1, value=log.id)
        ws.cell(row=row_idx, column=2, value=log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else "")
        ws.cell(row=row_idx, column=3, value=log.username or "")
        ws.cell(row=row_idx, column=4, value=log.action)
        ws.cell(row=row_idx, column=5, value=log.module)
        ws.cell(row=row_idx, column=6, value=log.description)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"audit_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Distinct values for filter dropdowns ──────────────────────

@router.get("/filters")
def get_filter_options(
    _admin: User = Depends(require_admin),
    db: Session  = Depends(get_db),
):
    """Return distinct action and module values for filter dropdowns."""
    actions = [r[0] for r in db.query(AuditLog.action).distinct().order_by(AuditLog.action).all()]
    modules = [r[0] for r in db.query(AuditLog.module).distinct().order_by(AuditLog.module).all()]
    return {"actions": actions, "modules": modules}
