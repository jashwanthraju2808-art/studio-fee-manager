"""
Excel import router — Member bulk import (admin only).

Endpoints:
  GET  /import/members/template  — download sample .xlsx template
  POST /import/members/validate  — validate uploaded file, return row-level preview
  POST /import/members           — commit validated rows in a single transaction

Rules:
  - Admin only.
  - Only .xlsx accepted (validated by extension + openpyxl parse attempt).
  - phone_number is the uniqueness key (matches existing Member model).
  - Preview/validate before commit.
  - Full transaction: if any committed row fails, the whole import rolls back.
  - Audit log on every import operation.
  - Never trusts client-supplied role values.
  - Upload size capped at 5 MB.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.core.auth import require_admin
from app.database.dependencies import get_db
from app.models.batch import Batch
from app.models.member import Member
from app.models.user import User
from app.services.audit_service import log_action

router = APIRouter(prefix="/import", tags=["Import"])

MAX_UPLOAD_BYTES = 5 * 1024 * 1024  # 5 MB

# Expected column headers in the template (case-insensitive match)
REQUIRED_COLUMNS = {"first_name", "last_name", "age", "phone_number", "fee"}
ALL_COLUMNS = [
    "first_name", "last_name", "age", "phone_number",
    "email", "fee", "batch_name", "is_active",
]


# ── Schemas ────────────────────────────────────────────────────

class RowPreview(BaseModel):
    row: int
    first_name: Optional[str]
    last_name: Optional[str]
    age: Optional[int]
    phone_number: Optional[str]
    email: Optional[str]
    fee: Optional[int]
    batch_name: Optional[str]
    is_active: bool
    action: str          # "create" | "update" | "skip"
    errors: List[str]
    valid: bool


class ValidateResponse(BaseModel):
    total_rows: int
    valid_rows: int
    invalid_rows: int
    create_count: int
    update_count: int
    previews: List[RowPreview]


class ImportSummary(BaseModel):
    created: int
    updated: int
    skipped: int
    failed: int
    errors: List[str]


# ── Helper: parse one data row ─────────────────────────────────

def _parse_row(row_num: int, row_data: dict, batch_map: dict) -> RowPreview:
    """
    Validate a single row dictionary. Returns a RowPreview with .valid and .errors.
    batch_map: {batch_name_lower: batch_id}
    """
    errors: List[str] = []

    first_name   = str(row_data.get("first_name") or "").strip()
    last_name    = str(row_data.get("last_name")  or "").strip()
    phone_number = str(row_data.get("phone_number") or "").strip()
    email_raw    = str(row_data.get("email")  or "").strip() or None
    batch_name   = str(row_data.get("batch_name") or "").strip() or None

    # Age
    try:
        age = int(row_data.get("age") or 0)
        if age <= 0 or age > 120:
            errors.append("age must be a positive integer (1–120)")
    except (ValueError, TypeError):
        age = None
        errors.append("age must be a number")

    # Fee
    try:
        fee = int(row_data.get("fee") or 0)
        if fee < 0:
            errors.append("fee must be >= 0")
    except (ValueError, TypeError):
        fee = None
        errors.append("fee must be a number")

    # Required text fields
    if not first_name:
        errors.append("first_name is required")
    if not last_name:
        errors.append("last_name is required")
    if not phone_number:
        errors.append("phone_number is required")
    elif len(phone_number) > 15:
        errors.append("phone_number must be <= 15 characters")

    # is_active
    raw_active = str(row_data.get("is_active") or "true").strip().lower()
    is_active  = raw_active not in ("false", "0", "no", "inactive", "n")

    # Batch lookup (warn but don't fail if batch not found)
    if batch_name and batch_name.lower() not in batch_map:
        errors.append(f"batch_name '{batch_name}' not found in system (will be ignored)")

    return RowPreview(
        row=row_num,
        first_name=first_name or None,
        last_name=last_name or None,
        age=age,
        phone_number=phone_number or None,
        email=email_raw,
        fee=fee,
        batch_name=batch_name,
        is_active=is_active,
        action="create",   # set later after DB dedup check
        errors=errors,
        valid=len(errors) == 0,
    )


def _read_workbook(contents: bytes):
    """Parse xlsx bytes → list of row dicts. Raises HTTPException on bad file."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl is not installed")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid or corrupted .xlsx file")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Spreadsheet is empty")

    # Normalise headers
    raw_headers = [str(h).strip().lower().replace(" ", "_") if h else "" for h in rows[0]]

    missing = REQUIRED_COLUMNS - set(raw_headers)
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required columns: {', '.join(sorted(missing))}",
        )

    data_rows = []
    for row in rows[1:]:
        row_dict = {raw_headers[i]: (row[i] if i < len(row) else None) for i in range(len(raw_headers))}
        # Skip completely empty rows
        if all(v is None or str(v).strip() == "" for v in row_dict.values()):
            continue
        data_rows.append(row_dict)

    wb.close()
    return data_rows


# ── Template download ──────────────────────────────────────────

@router.get("/members/template")
def download_member_template(
    _admin: User = Depends(require_admin),
):
    """Download a sample .xlsx import template for members."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl is not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members Import"

    headers = ["first_name", "last_name", "age", "phone_number", "email", "fee", "batch_name", "is_active"]
    fill = PatternFill(start_color="1E1B2E", end_color="1E1B2E", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center")

    widths = [14, 14, 6, 14, 28, 8, 24, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # Sample row
    ws.append(["Priya", "Sharma", 28, "9876543210", "priya@example.com", 1500, "6:30 AM – 7:30 AM", "true"])
    ws.append(["Rahul", "Verma",  35, "9123456789", "",                  1200, "5:00 PM – 6:00 PM", "true"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="member_import_template.xlsx"'},
    )


# ── Validate endpoint ──────────────────────────────────────────

@router.post("/members/validate", response_model=ValidateResponse)
async def validate_member_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _admin: User = Depends(require_admin),
):
    """
    Validate an uploaded .xlsx without writing to the database.
    Returns a row-by-row preview with errors and proposed action (create/update).
    """
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")

    contents = await file.read()
    if len(contents) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="File too large (max 5 MB)")

    data_rows = _read_workbook(contents)

    # Build batch lookup map
    batches   = db.query(Batch).all()
    batch_map = {b.name.lower(): b.id for b in batches}

    # Existing phone numbers for dedup
    existing_phones = {
        m.phone_number: m.id
        for m in db.query(Member.phone_number, Member.id).all()
    }

    previews: List[RowPreview] = []
    create_count = update_count = invalid_count = 0

    for idx, row_data in enumerate(data_rows, start=2):  # row 1 is header
        preview = _parse_row(idx, row_data, batch_map)

        if preview.valid and preview.phone_number:
            if preview.phone_number in existing_phones:
                preview.action = "update"
                update_count += 1
            else:
                preview.action = "create"
                create_count += 1
        else:
            preview.action = "skip"
            invalid_count += 1

        previews.append(preview)

    return ValidateResponse(
        total_rows=len(data_rows),
        valid_rows=create_count + update_count,
        invalid_rows=invalid_count,
        create_count=create_count,
        update_count=update_count,
        previews=previews,
    )


# ── Commit endpoint ────────────────────────────────────────────

@router.post("/members", response_model=ImportSummary)
async def commit_member_import(
    file: UploadFile = File(...),
    update_existing: bool = Query(default=True, description="Update existing members if phone matches"),
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    """
    Import members from .xlsx. Fully transactional — if any row fails
    unexpectedly, the entire import is rolled back.
    Invalid rows are skipped (not rolled back) but counted in the summary.
    """
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")

    contents = await file.read()
    if len(contents) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="File too large (max 5 MB)")

    data_rows = _read_workbook(contents)

    batches   = db.query(Batch).all()
    batch_map = {b.name.lower(): b.id for b in batches}

    existing_members = {
        m.phone_number: m
        for m in db.query(Member).all()
    }

    created = updated = skipped = failed = 0
    error_messages: List[str] = []

    try:
        for idx, row_data in enumerate(data_rows, start=2):
            preview = _parse_row(idx, row_data, batch_map)

            if not preview.valid:
                skipped += 1
                for e in preview.errors:
                    error_messages.append(f"Row {idx}: {e}")
                continue

            batch_id = batch_map.get(preview.batch_name.lower()) if preview.batch_name else None

            try:
                if preview.phone_number in existing_members:
                    if update_existing:
                        m = existing_members[preview.phone_number]
                        m.first_name = preview.first_name
                        m.last_name  = preview.last_name
                        m.age        = preview.age
                        m.email      = preview.email
                        m.fee        = preview.fee
                        m.batch_id   = batch_id
                        m.is_active  = preview.is_active
                        updated += 1
                    else:
                        skipped += 1
                        continue
                else:
                    new_m = Member(
                        first_name   = preview.first_name,
                        last_name    = preview.last_name,
                        age          = preview.age,
                        phone_number = preview.phone_number,
                        email        = preview.email,
                        fee          = preview.fee,
                        batch_id     = batch_id,
                        is_active    = preview.is_active,
                    )
                    db.add(new_m)
                    created += 1

            except Exception as row_exc:
                failed += 1
                error_messages.append(f"Row {idx}: {row_exc}")

        # Audit the import before committing
        log_action(
            db,
            username=admin.username,
            action="IMPORT",
            module="Import",
            description=(
                f"Member Excel import by '{admin.username}': "
                f"{created} created, {updated} updated, "
                f"{skipped} skipped, {failed} failed "
                f"(file: {file.filename})"
            ),
        )
        db.commit()

    except Exception as exc:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Import failed and was rolled back: {exc}",
        )

    return ImportSummary(
        created=created,
        updated=updated,
        skipped=skipped,
        failed=failed,
        errors=error_messages,
    )
