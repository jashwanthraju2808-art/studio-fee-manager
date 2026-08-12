"""
Excel export router.

Endpoints:
  GET /export/members        — any authenticated user
  GET /export/payments       — any authenticated user
  GET /export/attendance     — any authenticated user
  GET /export/users          — ADMIN ONLY (never exports hashed_password)
  GET /export/audit-logs     — ADMIN ONLY

Rules:
  - Never export hashed_password, JWT tokens, secrets, or credentials.
  - Empty datasets produce a valid .xlsx with headers only.
  - Filenames include a timestamp for uniqueness.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload

from app.core.auth import get_current_user, require_admin
from app.database.dependencies import get_db
from app.models.attendance import Attendance
from app.models.audit_log import AuditLog
from app.models.member import Member
from app.models.payment import Payment
from app.models.user import User

router = APIRouter(prefix="/export", tags=["Export"])

_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _make_response(wb, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type=_MIME,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _header_style():
    """Return (fill, font) for header rows."""
    from openpyxl.styles import Font, PatternFill
    fill = PatternFill(start_color="1E1B2E", end_color="1E1B2E", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    return fill, font


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w


def _ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


# ── Members ────────────────────────────────────────────────────

@router.get("/members")
def export_members(
    batch_id: Optional[int] = Query(default=None),
    active_only: bool = Query(default=True),
    db: Session = Depends(get_db),
    _current_user: User = Depends(get_current_user),
):
    """Export members to .xlsx. Any authenticated user."""
    import openpyxl
    from openpyxl.styles import Alignment

    q = db.query(Member).options(joinedload(Member.batch))
    if active_only:
        q = q.filter(Member.is_active == True)  # noqa: E712
    if batch_id:
        q = q.filter(Member.batch_id == batch_id)
    members = q.order_by(Member.first_name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members"

    headers = [
        "ID", "First Name", "Last Name", "Age", "Phone", "Email",
        "Fee (₹)", "Batch", "Active", "Joined",
    ]
    fill, font = _header_style()
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center")

    _set_col_widths(ws, [6, 16, 16, 6, 14, 28, 10, 22, 8, 20])

    for ri, m in enumerate(members, 2):
        ws.cell(ri, 1, m.id)
        ws.cell(ri, 2, m.first_name)
        ws.cell(ri, 3, m.last_name)
        ws.cell(ri, 4, m.age)
        ws.cell(ri, 5, m.phone_number)
        ws.cell(ri, 6, m.email or "")
        ws.cell(ri, 7, m.fee)
        ws.cell(ri, 8, m.batch.name if m.batch else "")
        ws.cell(ri, 9, "Yes" if m.is_active else "No")
        ws.cell(ri, 10, m.created_at.strftime("%Y-%m-%d") if m.created_at else "")

    return _make_response(wb, f"members_{_ts()}.xlsx")


# ── Payments ───────────────────────────────────────────────────

@router.get("/payments")
def export_payments(
    month: Optional[str] = Query(default=None, description="YYYY-MM"),
    db: Session = Depends(get_db),
    _current_user: User = Depends(get_current_user),
):
    """Export payments to .xlsx. Any authenticated user."""
    import openpyxl
    from openpyxl.styles import Alignment

    q = db.query(Payment).options(joinedload(Payment.member))
    if month:
        q = q.filter(Payment.month == month)
    payments = q.order_by(Payment.payment_date.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payments"

    headers = ["ID", "Member", "Amount (₹)", "Month", "Payment Date", "Note", "Recorded At"]
    fill, font = _header_style()
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center")

    _set_col_widths(ws, [6, 24, 12, 10, 14, 30, 20])

    for ri, p in enumerate(payments, 2):
        member_name = (
            f"{p.member.first_name} {p.member.last_name}" if p.member else f"id={p.member_id}"
        )
        ws.cell(ri, 1, p.id)
        ws.cell(ri, 2, member_name)
        ws.cell(ri, 3, p.amount)
        ws.cell(ri, 4, p.month)
        ws.cell(ri, 5, str(p.payment_date))
        ws.cell(ri, 6, p.note or "")
        ws.cell(ri, 7, p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else "")

    return _make_response(wb, f"payments_{month or 'all'}_{_ts()}.xlsx")


# ── Attendance ─────────────────────────────────────────────────

@router.get("/attendance")
def export_attendance(
    month: Optional[str] = Query(default=None, description="YYYY-MM"),
    db: Session = Depends(get_db),
    _current_user: User = Depends(get_current_user),
):
    """Export attendance to .xlsx. Any authenticated user."""
    import openpyxl
    from openpyxl.styles import Alignment

    q = db.query(Attendance).options(joinedload(Attendance.member))
    if month:
        try:
            year, mon = map(int, month.split("-"))
            from sqlalchemy import func
            q = q.filter(
                func.extract("year",  Attendance.att_date) == year,
                func.extract("month", Attendance.att_date) == mon,
            )
        except ValueError:
            pass
    records = q.order_by(Attendance.att_date.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    headers = ["ID", "Member", "Date", "Status"]
    fill, font = _header_style()
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center")

    _set_col_widths(ws, [6, 24, 14, 10])

    for ri, a in enumerate(records, 2):
        member_name = (
            f"{a.member.first_name} {a.member.last_name}" if a.member else f"id={a.member_id}"
        )
        ws.cell(ri, 1, a.id)
        ws.cell(ri, 2, member_name)
        ws.cell(ri, 3, str(a.att_date))
        ws.cell(ri, 4, "Present" if a.present else "Absent")

    return _make_response(wb, f"attendance_{month or 'all'}_{_ts()}.xlsx")


# ── Users (ADMIN ONLY — never exports hashed_password) ────────

@router.get("/users")
def export_users(
    db: Session = Depends(get_db),
    _admin: User = Depends(require_admin),
):
    """Export user list to .xlsx — ADMIN ONLY. Never includes hashed_password."""
    import openpyxl
    from openpyxl.styles import Alignment

    users = db.query(User).order_by(User.id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"

    headers = ["ID", "Username", "Role", "Active", "Created At"]
    fill, font = _header_style()
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center")

    _set_col_widths(ws, [6, 20, 12, 8, 20])

    for ri, u in enumerate(users, 2):
        ws.cell(ri, 1, u.id)
        ws.cell(ri, 2, u.username)
        ws.cell(ri, 3, u.role)
        ws.cell(ri, 4, "Yes" if u.is_active else "No")
        ws.cell(ri, 5, u.created_at.strftime("%Y-%m-%d %H:%M") if u.created_at else "")
        # hashed_password is intentionally NOT included

    return _make_response(wb, f"users_{_ts()}.xlsx")


# ── Audit Logs (ADMIN ONLY) ────────────────────────────────────

@router.get("/audit-logs")
def export_audit_logs_shortcut(
    db: Session = Depends(get_db),
    _admin: User = Depends(require_admin),
):
    """Convenience shortcut — redirects to /audit-logs/export logic inline."""
    import openpyxl
    from openpyxl.styles import Alignment

    logs = db.query(AuditLog).order_by(AuditLog.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Logs"

    headers = ["ID", "Date/Time", "Username", "Action", "Module", "Description"]
    fill, font = _header_style()
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center")

    _set_col_widths(ws, [6, 20, 18, 20, 16, 80])

    for ri, log in enumerate(logs, 2):
        ws.cell(ri, 1, log.id)
        ws.cell(ri, 2, log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else "")
        ws.cell(ri, 3, log.username or "")
        ws.cell(ri, 4, log.action)
        ws.cell(ri, 5, log.module)
        ws.cell(ri, 6, log.description)

    return _make_response(wb, f"audit_logs_{_ts()}.xlsx")
